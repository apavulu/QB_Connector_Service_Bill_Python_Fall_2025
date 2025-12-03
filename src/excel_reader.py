from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from typing import List
from src.models import BillRecord


def _parse_bank_date(raw_value) -> datetime | None:
    """
    Safely convert Excel date formats into datetime | None.
    Handles:
    - Excel datetime objects
    - Excel serial numbers
    - String dates
    """
    if raw_value is None or raw_value == "":
        return None

    # Case 1: Already a datetime
    if isinstance(raw_value, datetime):
        return raw_value

    # Case 2: Excel serial number (float or int)
    if isinstance(raw_value, (float, int)):
        try:
            # Excel base date is 1899-12-30
            return datetime.fromordinal(
                datetime(1899, 12, 30).toordinal() + int(raw_value)
            )
        except Exception:
            return None

    # Case 3: String date
    if isinstance(raw_value, str):
        raw_value = raw_value.strip()
        if not raw_value:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw_value, fmt)
            except ValueError:
                continue
        return None

    return None


def read_excel_data(file_path: Path) -> List[BillRecord]:
    """Read and parse Excel data dynamically based on column headers."""

    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty")

    header_row = [str(h).strip() if h else "" for h in rows[0]]
    header_index = {name: idx for idx, name in enumerate(header_row)}

    required_columns = [
        "Parent ID",
        "Child ID",
        "Supplier",
        "Check Amount",
        "Bank Date",
        "Tier 2 - Chart of Account",
    ]

    for col in required_columns:
        if col not in header_index:
            raise ValueError(f"Missing required column in Excel: '{col}'")

    bills: List[BillRecord] = []

    for row in rows[1:]:
        if not any(row):  # skip empty rows
            continue

        try:
            parent_id = str(row[header_index["Parent ID"]] or "").strip()
            child_id = str(row[header_index["Child ID"]] or "").strip()
            supplier = str(row[header_index["Supplier"]] or "").strip()

            memo = parent_id
            line_memo = child_id

            # Parse bank date safely
            raw_bank_date = row[header_index["Bank Date"]]
            bank_date = _parse_bank_date(raw_bank_date)

            chart_account = str(
                row[header_index["Tier 2 - Chart of Account"]] or ""
            ).strip()

            # Parse amount safely
            amount_value = row[header_index["Check Amount"]]
            if amount_value is None:
                amount = 0.0
            else:
                amount = float(
                    str(amount_value).replace("$", "").replace(",", "").strip()
                )

            record_id = parent_id
            if not record_id:
                continue

            bills.append(
                BillRecord(
                    record_id=record_id,
                    supplier=supplier,
                    bank_date=bank_date,
                    chart_account=chart_account,
                    amount=amount,
                    memo=memo,
                    line_memo=line_memo,
                    source="excel",
                )
            )

        except Exception as e:
            print(f"Skipping row due to error: {e}")
            continue

    print(f"Loaded {len(bills)} bill records from Excel.")
    return bills
